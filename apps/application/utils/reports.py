import io

from django.http import HttpResponse
from weasyprint import HTML
import openpyxl
from openpyxl.utils import get_column_letter


def generate_pdf_report(html_content: str) -> HttpResponse:
    """
    Generate PDF file from given HTML string and return as HTTP response.
    """
    pdf_file = HTML(string=html_content).write_pdf()
    response = HttpResponse(
        pdf_file,
        content_type="application/pdf"
    )
    response[
        'Content-Disposition'
    ] = 'attachment; filename="applications_report.pdf"'
    return response


def generate_xlsx_report(data: list, columns: list) -> HttpResponse:
    """
    Generate XLSX file from data and return as HTTP response.

    Args:
        data: List of rows (each row is a tuple or list of cell values).
        columns: List of column headers (strings).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications Report"

    # Write header row
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Set column widths
    for col_num in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Save to in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )
    response[
        'Content-Disposition'
    ] = 'attachment; filename="applications_report.xlsx"'
    return response
